"""Выгрузка данных для админки: кто зарегистрирован и кто с кем в комнате.

Два листа (в CSV — два файла, в JSON — два массива):
  * «Пользователи» — все анкеты с Telegram ID и никами;
  * «Комнаты» — состав комнат, по строке на каждого жильца: так удобно
    фильтровать и сводить в Excel.

Два режима: full — со всеми параметрами анкеты, short — только имя, ник и ID,
чтобы просто со всеми связаться. Фотографии не выгружаем: в таблице от них
толку нет.
"""

import csv
import io
import json
import zipfile
from datetime import datetime
from typing import List, Optional

import campuses
import models

# Подписи значений — те же, что человек видит на сайте. В выгрузке важнее
# читаемость, чем машинные коды: таблицу открывают глазами.
GENDER = {"male": "Парень", "female": "Девушка", "other": "Другое"}
TRACK = {
    "dev": "Разработка",
    "business": "Бизнес",
    "design": "Дизайн",
    "ai": "ИИ",
    "undecided": "Не определился",
}
SLEEP = {"lark": "Жаворонок", "owl": "Сова", "any": "Без разницы"}
SMOKING = {"no": "Не курит", "yes": "Курит", "vape": "Электронки"}
TIDINESS = {"relaxed": "Расслабленно", "medium": "Умеренно", "neat": "Аккуратно"}
WAKEUP = {
    "alarm_one": "Один будильник",
    "alarm_many": "Десять будильников",
    "natural": "Просыпается сам",
}
COOKING = {"self": "Сам", "together": "Вместе", "delivery": "Доставка"}
GUESTS = {"often": "Часто", "sometimes": "Иногда", "never": "Не зовёт"}
SHOWER = {"morning": "Утром", "evening": "Вечером", "any": "Когда как"}
TEMPERATURE = {"cool": "Прохладно", "medium": "Нормально", "warm": "Тепло"}
NOISE = {"quiet": "Тишина", "headphones": "В наушниках", "loud": "Музыка вслух"}
ALCOHOL = {"no": "Не пьёт", "sometimes": "Иногда", "often": "Часто"}

# «Не выбрано» показываем прочерком, а не пустой ячейкой: пустая читается как
# «данные потерялись», прочерк — как «человек не ответил».
UNSET = "—"

FULL = "full"
SHORT = "short"


def _label(mapping: dict, value: Optional[str]) -> str:
    return mapping.get(value or "", UNSET) or UNSET


def _yes_no(value: bool) -> str:
    return "да" if value else "нет"


def _cooking(value: str) -> str:
    items = [COOKING[key] for key in (value or "").split(",") if key in COOKING]
    return ", ".join(items) if items else UNSET


def _room_capacities(value: str) -> str:
    """«3 или 4» — желаемых размеров комнаты может быть несколько."""
    items = [part.strip() for part in (value or "").split(",") if part.strip()]
    if not items:
        return "Не важно"
    return " или ".join(items)


def _when(value: Optional[datetime]) -> str:
    return value.strftime("%d.%m.%Y %H:%M") if value else ""


def _link(username: str) -> str:
    return f"https://t.me/{username}" if username else ""


def users_table(profiles: List[models.Profile], scope: str) -> tuple[list, list]:
    """Лист «Пользователи»: заголовки и строки."""
    if scope == SHORT:
        headers = ["Имя", "Юзернейм", "Telegram ID", "Ссылка"]
        rows = [
            [
                p.name,
                f"@{p.telegram}" if p.telegram else UNSET,
                p.telegram_id or "",
                _link(p.telegram),
            ]
            for p in profiles
        ]
        return headers, rows

    headers = [
        "ID",
        "Имя",
        "Юзернейм",
        "Telegram ID",
        "Ссылка",
        "Подтверждён",
        "Бот подключён",
        "Кампус-отель",
        "Пол",
        "Курс",
        "Направление",
        "Хочет комнату на",
        "Комната №",
        "Блок №",
        "Соседи",
        "Режим сна",
        "Подъём",
        "Душ",
        "Аккуратность",
        "Температура",
        "Звук",
        "Готовка",
        "Гости",
        "Курение",
        "Алкоголь",
        "О себе",
        "Анкета создана",
    ]
    rows = []
    for p in profiles:
        mates = (
            ", ".join(m.name for m in p.group.members if m.id != p.id)
            if p.group_id and p.group
            else ""
        )
        rows.append(
            [
                p.id,
                p.name,
                f"@{p.telegram}" if p.telegram else UNSET,
                p.telegram_id or "",
                _link(p.telegram),
                _yes_no(p.telegram_verified),
                # Уведомления доходят только до тех, кто нажал /start у бота.
                _yes_no(bool(p.telegram_chat_id)),
                campuses.label(p.campus),
                _label(GENDER, p.gender),
                p.course or UNSET,
                _label(TRACK, p.track),
                _room_capacities(p.room_capacities),
                p.group_id or "",
                (p.group.block_id if p.group_id and p.group else None) or "",
                mates or UNSET,
                _label(SLEEP, p.sleep_schedule),
                _label(WAKEUP, p.wakeup),
                _label(SHOWER, p.shower),
                _label(TIDINESS, p.tidiness),
                _label(TEMPERATURE, p.temperature),
                _label(NOISE, p.noise),
                _cooking(p.cooking),
                _label(GUESTS, p.guests),
                _label(SMOKING, p.smoking),
                _label(ALCOHOL, p.alcohol),
                (p.bio or "").replace("\n", " ").strip() or UNSET,
                _when(p.created_at),
            ]
        )
    return headers, rows


def rooms_table(groups: List[models.Group], scope: str) -> tuple[list, list]:
    """Лист «Комнаты»: по строке на жильца — так удобнее фильтровать.

    Комнату видно целиком: номер повторяется в каждой её строке.
    """
    ordered = sorted(groups, key=lambda g: (g.campus, -len(g.members), g.id))

    if scope == SHORT:
        headers = ["Комната №", "Участник", "Юзернейм"]
        rows = []
        for group in ordered:
            for member in group.members:
                rows.append(
                    [
                        group.id,
                        member.name,
                        f"@{member.telegram}" if member.telegram else UNSET,
                    ]
                )
        return headers, rows

    headers = [
        "Комната №",
        "Блок №",
        "Соседи по блоку",
        "Кампус-отель",
        "Пол",
        "Размер",
        "Занято",
        "Свободно",
        "Собрана",
        "Участник",
        "Юзернейм",
        "Telegram ID",
        "Ссылка",
        "Курс",
        "Направление",
        "Комната создана",
    ]
    rows = []
    for group in ordered:
        taken = len(group.members)
        # Соседи по блоку — вторая комната блока: с ней делят общий вход.
        neighbours = (
            ", ".join(
                m.name
                for other in group.block.groups
                if other.id != group.id
                for m in other.members
            )
            if group.block_id and group.block
            else ""
        )
        for member in group.members:
            rows.append(
                [
                    group.id,
                    group.block_id or "",
                    neighbours or UNSET,
                    campuses.label(group.campus),
                    _label(GENDER, group.gender),
                    group.capacity,
                    taken,
                    group.capacity - taken,
                    _yes_no(group.capacity - taken <= 0),
                    member.name,
                    f"@{member.telegram}" if member.telegram else UNSET,
                    member.telegram_id or "",
                    _link(member.telegram),
                    member.course or UNSET,
                    _label(TRACK, member.track),
                    _when(group.created_at),
                ]
            )
    return headers, rows


def _sheets(
    profiles: List[models.Profile], groups: List[models.Group], scope: str
) -> list[tuple[str, list, list]]:
    return [
        ("Пользователи", *users_table(profiles, scope)),
        ("Комнаты", *rooms_table(groups, scope)),
    ]


def to_xlsx(
    profiles: List[models.Profile], groups: List[models.Group], scope: str
) -> bytes:
    """Книга Excel: пользователи на первом листе, комнаты на втором."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    book = Workbook()
    book.remove(book.active)  # вместо листа по умолчанию делаем свои

    for title, headers, rows in _sheets(profiles, groups, scope):
        sheet = book.create_sheet(title)
        sheet.append(headers)
        for row in rows:
            sheet.append(row)

        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(vertical="center")
        # Шапка остаётся на месте, и по ней можно фильтровать.
        sheet.freeze_panes = "A2"
        if rows:
            sheet.auto_filter.ref = (
                f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
            )

        # Ширина по содержимому, но без «простыней» на поле «О себе».
        for index, header in enumerate(headers, start=1):
            widths = [len(str(header))] + [
                len(str(row[index - 1])) for row in rows
            ]
            sheet.column_dimensions[get_column_letter(index)].width = min(
                max(max(widths) + 2, 10), 45
            )

    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def to_csv_zip(
    profiles: List[models.Profile], groups: List[models.Group], scope: str
) -> bytes:
    """Два CSV в одном архиве: у CSV нет листов, а таблицы всё равно две.

    Разделитель «;» и BOM — чтобы файл открывался в Excel двойным кликом,
    а кириллица не превращалась в кракозябры.
    """
    archive = io.BytesIO()
    with zipfile.ZipFile(archive, "w", zipfile.ZIP_DEFLATED) as zf:
        for name, (title, headers, rows) in zip(
            ("users.csv", "rooms.csv"), _sheets(profiles, groups, scope)
        ):
            text = io.StringIO()
            writer = csv.writer(text, delimiter=";", lineterminator="\r\n")
            writer.writerow(headers)
            writer.writerows(rows)
            zf.writestr(name, "﻿" + text.getvalue())
    return archive.getvalue()


def to_json(
    profiles: List[models.Profile], groups: List[models.Group], scope: str
) -> bytes:
    """JSON — для тех случаев, когда данные нужно обработать программой."""
    payload = {"exported_at": datetime.utcnow().isoformat(timespec="seconds")}
    for key, (title, headers, rows) in zip(
        ("users", "rooms"), _sheets(profiles, groups, scope)
    ):
        payload[key] = [dict(zip(headers, row)) for row in rows]
    return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")
